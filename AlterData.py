import tkinter as tk
import openpyxl

def save_register_details():
    # Get the entered values from the Entry widgets
    name = name_entry.get()
    age = age_entry.get()
    email = email_entry.get()

    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook('register_details.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet (the first sheet by default)
    sheet = workbook.active

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the register details to the Excel sheet
    sheet.cell(row=next_row, column=1).value = name
    sheet.cell(row=next_row, column=2).value = age
    sheet.cell(row=next_row, column=3).value = email

    # Save the changes to the Excel file
    workbook.save('register_details.xlsx')

    # Clear the Entry widgets
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)

# Create the main window
window = tk.Tk()
window.title("Register Details")

# Create labels and Entry widgets for each detail
name_label = tk.Label(window, text="Name:")
name_label.pack()
name_entry = tk.Entry(window)
name_entry.pack()

age_label = tk.Label(window, text="Age:")
age_label.pack()
age_entry = tk.Entry(window)
age_entry.pack()

email_label = tk.Label(window, text="Email:")
email_label.pack()
email_entry = tk.Entry(window)
email_entry.pack()

# Create the save button
save_button = tk.Button(window, text="Save", command=save_register_details)
save_button.pack()

# Run the Tkinter event loop
window.mainloop()
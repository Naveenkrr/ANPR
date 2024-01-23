import openpyxl



def isAllowed(vNumber):
    workbook = openpyxl.load_workbook('Registration_info.xlsx')
    sheet = workbook['Sheet1']
    for row in sheet.iter_rows(min_row=1 , max_row=sheet.max_row , min_col=6, max_col=6 , values_only=True):
        if vNumber in row:
            return True
    return False

print(isAllowed('UP 32 JJ 7591'))
